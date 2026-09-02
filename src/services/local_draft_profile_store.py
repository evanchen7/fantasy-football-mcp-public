"""Sanitized private storage and DraftSheets conversion for local draft profiles.

The recommendation engine can use this profile while Yahoo API access is unavailable.
Workbook parsing is deliberately narrow: only the DraftSheets ``ECR`` and ``Scoring``
cells needed by the normalized schema are read, and the workbook itself is never stored.
"""

from __future__ import annotations

import hashlib
import io
import json
import math
import os
import re
import tempfile
import threading
import unicodedata
import zipfile
from collections.abc import Mapping, Sequence
from copy import deepcopy
from datetime import date, datetime, timezone
from pathlib import Path, PurePosixPath
from typing import Any

from src.services.yahoo_player_identity import normalize_yahoo_player_key

DEFAULT_PROFILE_STORE_PATH = Path.home() / ".fantasy-football-mcp" / "draft-profiles.json"
DEFAULT_PROFILE_DEFAULTS_STORE_PATH = (
    Path.home() / ".fantasy-football-mcp" / "draft-profile-defaults.json"
)
MAX_CANDIDATES = 500
MAX_PROFILE_DEFAULTS = 32
MAX_ECR_ROWS = 2_000
MAX_SCORING_ROWS = 100
MAX_XLSX_BYTES = 8 * 1024 * 1024
MAX_XLSX_MEMBERS = 128
MAX_XLSX_UNCOMPRESSED_BYTES = 16 * 1024 * 1024

_STORE_LOCK = threading.Lock()
_SESSION_KEY = re.compile(r"^[A-Za-z0-9_-]{1,32}:[A-Za-z0-9_-]{1,64}$")
_IDENTIFIER = re.compile(r"^[A-Za-z0-9_-]+$")
_TEAM = re.compile(r"^[A-Z0-9]{1,8}$")
_CANDIDATE_POSITIONS = {"QB", "RB", "WR", "TE", "K", "DST"}
_ROSTER_ORDER = (
    "QB",
    "RB",
    "WR",
    "TE",
    "FLEX",
    "SUPERFLEX",
    "K",
    "DST",
    "BN",
    "IR",
)
_ROSTER_POSITIONS = set(_ROSTER_ORDER)
_POSITION_ALIASES = {
    "DEF": "DST",
    "D/ST": "DST",
    "D-ST": "DST",
    "W/R/T": "FLEX",
    "W-R-T": "FLEX",
    "Q/W/R/T": "SUPERFLEX",
    "Q-W-R-T": "SUPERFLEX",
    "OP": "SUPERFLEX",
    "BE": "BN",
    "BENCH": "BN",
}
_PROVENANCE_FORMATS = {"draftsheets-2026", "csv", "json"}
_ECR_FIELD_ALIASES = {
    "rank": {"RK", "RANK", "OVERALLRANK"},
    "name": {"PLAYERNAME", "PLAYER", "NAME"},
    "team": {"TEAM", "NFLTEAM"},
    "position": {"POS", "POSITION"},
    "adp": {"ADP", "AVERAGEDRAFTPOSITION"},
    "bye": {"BYE", "BYEWEEK"},
    "player_key": {"PLAYERKEY", "YAHOOPLAYERKEY"},
}
_SCORING_ALIASES = {
    "TEAMS": "teams",
    "TEAMCOUNT": "teams",
    "NUMTEAMS": "teams",
    "NUMBEROFTEAMS": "teams",
    "QB": "QB",
    "RB": "RB",
    "WR": "WR",
    "TE": "TE",
    "FLEX": "FLEX",
    "WRT": "FLEX",
    "SUPERFLEX": "SUPERFLEX",
    "QWRT": "SUPERFLEX",
    "OP": "SUPERFLEX",
    "K": "K",
    "KICKER": "K",
    "DST": "DST",
    "DEF": "DST",
    "DEFENSE": "DST",
    "BENCH": "BN",
    "BN": "BN",
    "BE": "BN",
    "IR": "IR",
}
_PRIVATE_PLAYER_TEXT = re.compile(
    r"(?:[a-z][a-z0-9+.-]{1,15}://|www\.|"
    r"\b(?:[a-z0-9-]+\.)+(?:com|net|org|io|co|dev|app|test)(?:[/:?#]|$)|"
    r"(?:^|[?&;\s])"
    r"(?:auth(?:orization)?|token|api[_-]?key|key|cookie|session|password|secret)\s*[:=]|"
    r"\?[^\s=]{1,64}=)",
    re.IGNORECASE,
)
_MACRO_PART_PREFIXES = (
    "xl/macrosheets/",
    "xl/dialogsheets/",
    "xl/activex/",
    "xl/ctrlprops/",
    "customui/",
)
_MACRO_PACKAGE_MARKERS = (
    b"vbaproject",
    b"macrosheet",
    b"intlmacrosheet",
    b"macroenabled",
)


class LocalDraftProfileValidationError(ValueError):
    """Raised when imported local draft-profile data is unsafe or invalid."""


class LocalDraftProfileNotFoundError(LocalDraftProfileValidationError):
    """Raised when an explicitly selected reusable profile is unavailable."""


class LocalDraftProfileConflictError(LocalDraftProfileValidationError):
    """Raised when an explicit profile bind would cross or replace an identity."""


def _profile_store_path(path: str | Path | None = None) -> Path:
    configured = (
        path or os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_PATH") or DEFAULT_PROFILE_STORE_PATH
    )
    return Path(configured).expanduser()


def _profile_defaults_store_path(
    path: str | Path | None = None,
    *,
    profile_path: str | Path | None = None,
) -> Path:
    configured = path or os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_DEFAULTS_PATH")
    if configured:
        return Path(configured).expanduser()
    if profile_path is not None or os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_PATH"):
        return _profile_store_path(profile_path).with_name("draft-profile-defaults.json")
    return DEFAULT_PROFILE_DEFAULTS_STORE_PATH


def _safe_string(value: Any, field: str, maximum: int = 100) -> str:
    if not isinstance(value, str):
        raise LocalDraftProfileValidationError(f"{field} must be a non-empty string")
    normalized = unicodedata.normalize("NFKC", value)
    result = " ".join(normalized.split())
    if not result:
        raise LocalDraftProfileValidationError(f"{field} must be a non-empty string")
    if len(result) > maximum:
        raise LocalDraftProfileValidationError(f"{field} is too long")
    return result


def _safe_player_name(value: Any, field: str) -> str:
    if not isinstance(value, str):
        raise LocalDraftProfileValidationError("player name is invalid")
    normalized = unicodedata.normalize("NFKC", value)
    if any(unicodedata.category(character) in {"Cc", "Cf", "Cs"} for character in normalized):
        raise LocalDraftProfileValidationError("player name is invalid")
    result = _safe_string(normalized, field)
    if result[0] in "=+@-" or _PRIVATE_PLAYER_TEXT.search(result):
        raise LocalDraftProfileValidationError("player name is invalid")
    return result


def _strict_integer(value: Any, field: str, minimum: int, maximum: int) -> int:
    if isinstance(value, bool) or not isinstance(value, int):
        raise LocalDraftProfileValidationError(f"{field} must be an integer")
    if value < minimum or value > maximum:
        raise LocalDraftProfileValidationError(f"{field} must be between {minimum} and {maximum}")
    return value


def _strict_number(value: Any, field: str, minimum: float, maximum: float) -> float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise LocalDraftProfileValidationError(f"{field} must be a number")
    result = float(value)
    if not math.isfinite(result) or result < minimum or result > maximum:
        raise LocalDraftProfileValidationError(
            f"{field} must be between {minimum:g} and {maximum:g}"
        )
    return result


def _coerce_integer(value: Any, field: str, minimum: int, maximum: int) -> int:
    if isinstance(value, bool):
        raise LocalDraftProfileValidationError(f"{field} must be an integer")
    try:
        number = float(value)
    except (TypeError, ValueError) as error:
        raise LocalDraftProfileValidationError(f"{field} must be an integer") from error
    if not math.isfinite(number) or not number.is_integer():
        raise LocalDraftProfileValidationError(f"{field} must be an integer")
    return _strict_integer(int(number), field, minimum, maximum)


def _coerce_number(value: Any, field: str, minimum: float, maximum: float) -> float:
    if isinstance(value, bool):
        raise LocalDraftProfileValidationError(f"{field} must be a number")
    try:
        number = float(value)
    except (TypeError, ValueError) as error:
        raise LocalDraftProfileValidationError(f"{field} must be a number") from error
    return _strict_number(number, field, minimum, maximum)


def _normalize_timestamp(value: Any, field: str) -> str:
    text = _safe_string(value, field, 64)
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except ValueError as error:
        raise LocalDraftProfileValidationError(f"{field} must be an ISO-8601 timestamp") from error
    if parsed.tzinfo is None or parsed.utcoffset() is None:
        raise LocalDraftProfileValidationError(f"{field} must include a timezone offset")
    result = parsed.astimezone(timezone.utc).isoformat()
    if result.endswith("+00:00"):
        result = f"{result[:-6]}Z"
    return result


def _parse_timestamp(value: str) -> datetime:
    return datetime.fromisoformat(value.replace("Z", "+00:00"))


def _resolve_current_season(current_season: int | None) -> int:
    if current_season is None:
        return datetime.now(timezone.utc).year
    return _strict_integer(current_season, "current_season", 2020, 2100)


def _require_current_default_profile_season(
    profile: Mapping[str, Any], current_season: int | None
) -> None:
    expected = _resolve_current_season(current_season)
    actual = profile["season"]
    if actual != expected:
        raise LocalDraftProfileConflictError(
            f"selected local draft profile is for season {actual}, not current UTC "
            f"season {expected}; choose a current-season profile or clear this default"
        )


def _sanitize_draft_identity(value: Any) -> dict[str, str]:
    if not isinstance(value, Mapping):
        raise LocalDraftProfileValidationError("draft identity is required")
    sport = _safe_string(value.get("sport"), "draft.sport", 32)
    league_id = _safe_string(value.get("leagueId"), "draft.leagueId", 64)
    team_id = _safe_string(value.get("teamId"), "draft.teamId", 64)
    for field, item in (
        ("draft.sport", sport),
        ("draft.leagueId", league_id),
        ("draft.teamId", team_id),
    ):
        if not _IDENTIFIER.fullmatch(item):
            raise LocalDraftProfileValidationError(f"{field} has an invalid format")
    session_key = _safe_string(value.get("sessionKey"), "draft.sessionKey", 100)
    if not _SESSION_KEY.fullmatch(session_key):
        raise LocalDraftProfileValidationError("draft.sessionKey has an invalid format")
    if session_key != f"{sport}:{league_id}":
        raise LocalDraftProfileValidationError("draft.sessionKey must equal sport:leagueId")
    return {
        "sport": sport,
        "leagueId": league_id,
        "teamId": team_id,
        "sessionKey": session_key,
    }


def _normalize_position(value: Any, *, roster: bool) -> str:
    text = _safe_string(value, "roster position" if roster else "ranking position", 20)
    position = re.sub(r"\s+", "", text.upper())
    if not roster:
        position = re.sub(r"\d+$", "", position)
    position = _POSITION_ALIASES.get(position, position)
    allowed = _ROSTER_POSITIONS if roster else _CANDIDATE_POSITIONS
    if position not in allowed:
        label = "roster position" if roster else "ranking position"
        raise LocalDraftProfileValidationError(f"unsupported {label}: {text}")
    return position


def _sanitize_candidate(value: Any, index: int) -> dict[str, Any]:
    if not isinstance(value, Mapping):
        raise LocalDraftProfileValidationError(f"rankings[{index}] must be an object")
    name = _safe_player_name(value.get("name"), f"rankings[{index}].name")
    position = _normalize_position(value.get("position"), roster=False)
    rank = _strict_integer(value.get("rank"), f"rankings[{index}].rank", 1, 10_000)
    result: dict[str, Any] = {
        "name": name,
        "position": position,
        "rank": rank,
    }
    if not _is_blank(value.get("team")):
        team = _safe_string(value.get("team"), f"rankings[{index}].team", 8).upper()
        if not _TEAM.fullmatch(team):
            raise LocalDraftProfileValidationError(f"rankings[{index}].team has an invalid format")
        result["team"] = team
    if "average_draft_position" in value and value["average_draft_position"] is not None:
        result["average_draft_position"] = _strict_number(
            value["average_draft_position"],
            f"rankings[{index}].average_draft_position",
            0.01,
            10_000,
        )
    if "bye_week" in value and value["bye_week"] is not None:
        result["bye_week"] = _strict_integer(
            value["bye_week"], f"rankings[{index}].bye_week", 1, 22
        )
    if "player_key" in value and value["player_key"] is not None:
        player_key = normalize_yahoo_player_key(value["player_key"])
        if player_key is None:
            raise LocalDraftProfileValidationError(
                f"rankings[{index}].player_key has an invalid format"
            )
        result["player_key"] = player_key
    return result


def _sanitize_league_settings(value: Any) -> dict[str, Any]:
    if not isinstance(value, Mapping):
        raise LocalDraftProfileValidationError("leagueSettings is required")
    teams = _strict_integer(value.get("teams"), "leagueSettings.teams", 2, 20)
    raw_positions = value.get("rosterPositions")
    if not isinstance(raw_positions, list) or not raw_positions:
        raise LocalDraftProfileValidationError(
            "leagueSettings.rosterPositions must be a non-empty list"
        )
    if len(raw_positions) > len(_ROSTER_POSITIONS):
        raise LocalDraftProfileValidationError("too many roster positions")
    positions: dict[str, int] = {}
    for index, item in enumerate(raw_positions):
        if not isinstance(item, Mapping):
            raise LocalDraftProfileValidationError(f"rosterPositions[{index}] must be an object")
        position = _normalize_position(item.get("position"), roster=True)
        if position in positions:
            raise LocalDraftProfileValidationError(f"duplicate roster position: {position}")
        positions[position] = _strict_integer(
            item.get("count"), f"rosterPositions[{index}].count", 1, 30
        )
    if sum(positions.values()) > 40:
        raise LocalDraftProfileValidationError("roster cannot exceed 40 slots")
    return {
        "teams": teams,
        "rosterPositions": [
            {"position": position, "count": positions[position]}
            for position in _ROSTER_ORDER
            if position in positions
        ],
    }


def _sanitize_provenance(value: Any, season: int) -> dict[str, str]:
    if not isinstance(value, Mapping):
        raise LocalDraftProfileValidationError("provenance is required")
    if value.get("kind") != "user-import":
        raise LocalDraftProfileValidationError("provenance.kind must be user-import")
    format_name = value.get("format")
    if format_name not in _PROVENANCE_FORMATS:
        raise LocalDraftProfileValidationError("unsupported provenance.format")
    result = {"kind": "user-import", "format": str(format_name)}
    if value.get("asOf") not in (None, ""):
        as_of = _safe_string(value.get("asOf"), "provenance.asOf", 10)
        try:
            parsed = date.fromisoformat(as_of)
        except ValueError as error:
            raise LocalDraftProfileValidationError("provenance.asOf must be an ISO date") from error
        if parsed.year != season:
            raise LocalDraftProfileValidationError("provenance asOf year must match season")
        result["asOf"] = parsed.isoformat()
    return result


def sanitize_local_draft_profile(value: Any) -> dict[str, Any]:
    """Return a strict allowlisted local profile safe for private persistence."""

    if not isinstance(value, Mapping) or value.get("schemaVersion") != 1:
        raise LocalDraftProfileValidationError("schemaVersion 1 is required")
    if value.get("source") != "local-draft-profile":
        raise LocalDraftProfileValidationError("unsupported local draft profile source")
    season = _strict_integer(value.get("season"), "season", 2020, 2100)
    imported_at = _normalize_timestamp(value.get("importedAt"), "importedAt")
    draft = _sanitize_draft_identity(value.get("draft"))
    raw_rankings = value.get("rankings")
    if not isinstance(raw_rankings, list) or not raw_rankings:
        raise LocalDraftProfileValidationError("rankings must be a non-empty list")
    if len(raw_rankings) > MAX_CANDIDATES:
        raise LocalDraftProfileValidationError(
            f"rankings cannot exceed {MAX_CANDIDATES} candidates"
        )
    rankings = [
        _sanitize_candidate(candidate, index) for index, candidate in enumerate(raw_rankings)
    ]
    rankings.sort(key=lambda candidate: (candidate["rank"], candidate["name"]))
    ranks: set[int] = set()
    identities: set[tuple[str, str, str]] = set()
    for candidate in rankings:
        rank = candidate["rank"]
        if rank in ranks:
            raise LocalDraftProfileValidationError(f"duplicate rank: {rank}")
        ranks.add(rank)
        identity = (
            candidate["name"].casefold(),
            candidate["position"],
            candidate.get("team", ""),
        )
        if identity in identities:
            raise LocalDraftProfileValidationError(f"duplicate player: {candidate['name']}")
        identities.add(identity)
    league_settings = _sanitize_league_settings(value.get("leagueSettings"))
    provenance = _sanitize_provenance(value.get("provenance"), season)
    return {
        "schemaVersion": 1,
        "source": "local-draft-profile",
        "season": season,
        "importedAt": imported_at,
        "draft": draft,
        "rankings": rankings,
        "leagueSettings": league_settings,
        "provenance": provenance,
    }


def _read_all(path: Path) -> dict[str, dict[str, Any]]:
    if path.is_symlink():
        raise LocalDraftProfileValidationError(
            "local draft profile store cannot be a symbolic link"
        )
    if not path.exists():
        return {}
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise LocalDraftProfileValidationError(
            f"could not read local draft profile store: {error}"
        ) from error
    if not isinstance(raw, dict):
        raise LocalDraftProfileValidationError("local draft profile store is malformed")
    sessions: dict[str, dict[str, Any]] = {}
    for key, value in raw.items():
        if not isinstance(key, str):
            raise LocalDraftProfileValidationError("local draft profile store is malformed")
        try:
            sanitized = sanitize_local_draft_profile(value)
        except LocalDraftProfileValidationError as error:
            raise LocalDraftProfileValidationError(
                f"local draft profile store is malformed: {error}"
            ) from error
        if sanitized["draft"]["sessionKey"] != key:
            raise LocalDraftProfileValidationError(
                "local draft profile store is malformed: session key mismatch"
            )
        sessions[key] = sanitized
    return sessions


def _prepare_store_directory(destination: Path, *, tighten_existing: bool) -> None:
    parent = destination.parent
    try:
        parent.mkdir(mode=0o700, parents=True, exist_ok=False)
        created = True
    except FileExistsError as error:
        if not parent.is_dir():
            raise LocalDraftProfileValidationError(
                "local draft profile parent is not a directory"
            ) from error
        created = False
    if created:
        parent.chmod(0o700)
    elif tighten_existing:
        if parent.is_symlink():
            raise LocalDraftProfileValidationError(
                "default local draft profile directory cannot be a symbolic link"
            )
        parent.chmod(0o700)


def _write_all(sessions: Mapping[str, Any], destination: Path) -> None:
    handle, temporary_name = tempfile.mkstemp(
        prefix=".draft-profiles-", suffix=".json", dir=destination.parent
    )
    try:
        with os.fdopen(handle, "w", encoding="utf-8") as temporary:
            json.dump(sessions, temporary, indent=2, sort_keys=True, allow_nan=False)
            temporary.write("\n")
            temporary.flush()
            os.fsync(temporary.fileno())
        os.chmod(temporary_name, 0o600)
        os.replace(temporary_name, destination)
    finally:
        if os.path.exists(temporary_name):
            os.unlink(temporary_name)


def _sanitize_profile_default(value: Any) -> dict[str, str]:
    if not isinstance(value, Mapping) or set(value) != {"sport", "sourceLeagueId"}:
        raise LocalDraftProfileValidationError(
            "local draft profile default fields are invalid"
        )
    sport = _safe_string(value.get("sport"), "sport", 32)
    source_league_id = _safe_string(
        value.get("sourceLeagueId"), "sourceLeagueId", 64
    )
    if not _IDENTIFIER.fullmatch(sport):
        raise LocalDraftProfileValidationError("sport has an invalid format")
    if not _IDENTIFIER.fullmatch(source_league_id):
        raise LocalDraftProfileValidationError(
            "sourceLeagueId has an invalid format"
        )
    return {"sport": sport, "sourceLeagueId": source_league_id}


def _read_profile_defaults(path: Path) -> list[dict[str, str]]:
    if path.is_symlink():
        raise LocalDraftProfileValidationError(
            "local draft profile default store cannot be a symbolic link"
        )
    if not path.exists():
        return []
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as error:
        raise LocalDraftProfileValidationError(
            f"could not read local draft profile default store: {error}"
        ) from error
    if (
        not isinstance(raw, dict)
        or set(raw) != {"schemaVersion", "defaults"}
        or raw.get("schemaVersion") != 1
        or isinstance(raw.get("schemaVersion"), bool)
        or not isinstance(raw.get("defaults"), list)
        or len(raw["defaults"]) > MAX_PROFILE_DEFAULTS
    ):
        raise LocalDraftProfileValidationError(
            "local draft profile default store is malformed"
        )
    defaults = [_sanitize_profile_default(value) for value in raw["defaults"]]
    sports = [value["sport"] for value in defaults]
    if len(sports) != len(set(sports)):
        raise LocalDraftProfileValidationError(
            "local draft profile default store contains duplicate sports"
        )
    return sorted(defaults, key=lambda value: value["sport"])


def _write_profile_defaults(
    defaults: Sequence[Mapping[str, str]], destination: Path
) -> None:
    handle, temporary_name = tempfile.mkstemp(
        prefix=".draft-profile-defaults-", suffix=".json", dir=destination.parent
    )
    try:
        with os.fdopen(handle, "w", encoding="utf-8") as temporary:
            json.dump(
                {"schemaVersion": 1, "defaults": list(defaults)},
                temporary,
                indent=2,
                sort_keys=True,
                allow_nan=False,
            )
            temporary.write("\n")
            temporary.flush()
            os.fsync(temporary.fileno())
        os.chmod(temporary_name, 0o600)
        os.replace(temporary_name, destination)
    finally:
        if os.path.exists(temporary_name):
            os.unlink(temporary_name)


def list_local_draft_profile_defaults(
    path: str | Path | None = None,
) -> list[dict[str, str]]:
    """List privacy-minimal per-sport default profile pointers."""

    custom_path = path is not None or bool(
        os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_DEFAULTS_PATH")
    ) or bool(os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_PATH"))
    destination = _profile_defaults_store_path(path)
    with _STORE_LOCK:
        if not custom_path and destination.parent.is_symlink():
            raise LocalDraftProfileValidationError(
                "default local draft profile directory cannot be a symbolic link"
            )
        defaults = _read_profile_defaults(destination)
    return defaults


def load_default_local_draft_profile(
    sport: str,
    path: str | Path | None = None,
) -> dict[str, str] | None:
    """Load only the explicit default pointer for one exact sport."""

    clean_sport = _safe_string(sport, "sport", 32)
    if not _IDENTIFIER.fullmatch(clean_sport):
        raise LocalDraftProfileValidationError("sport has an invalid format")
    return next(
        (
            value
            for value in list_local_draft_profile_defaults(path)
            if value["sport"] == clean_sport
        ),
        None,
    )


def set_default_local_draft_profile(
    sport: str,
    source_league_id: str,
    *,
    profile_path: str | Path | None = None,
    defaults_path: str | Path | None = None,
    current_season: int | None = None,
) -> dict[str, str]:
    """Atomically select an existing same-sport profile as the default."""

    selected = _sanitize_profile_default(
        {"sport": sport, "sourceLeagueId": source_league_id}
    )
    profile_destination = _profile_store_path(profile_path)
    defaults_destination = _profile_defaults_store_path(
        defaults_path, profile_path=profile_path
    )
    custom_profile_path = profile_path is not None or bool(
        os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_PATH")
    )
    custom_defaults_path = (
        defaults_path is not None
        or bool(os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_DEFAULTS_PATH"))
        or profile_path is not None
        or bool(os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_PATH"))
    )
    with _STORE_LOCK:
        if not custom_profile_path and profile_destination.parent.is_symlink():
            raise LocalDraftProfileValidationError(
                "default local draft profile directory cannot be a symbolic link"
            )
        profiles = _read_all(profile_destination)
        source_key = f"{selected['sport']}:{selected['sourceLeagueId']}"
        source = profiles.get(source_key)
        if source is None:
            source_sports = {
                profile["draft"]["sport"]
                for profile in profiles.values()
                if profile["draft"]["leagueId"] == selected["sourceLeagueId"]
            }
            if source_sports:
                raise LocalDraftProfileConflictError(
                    "selected local profile belongs to a different sport"
                )
            raise LocalDraftProfileNotFoundError(
                "selected local draft profile was not found"
            )
        _require_current_default_profile_season(source, current_season)
        _prepare_store_directory(
            defaults_destination, tighten_existing=not custom_defaults_path
        )
        defaults = _read_profile_defaults(defaults_destination)
        existing = next(
            (value for value in defaults if value["sport"] == selected["sport"]),
            None,
        )
        if existing == selected:
            defaults_destination.chmod(0o600)
            return existing
        updated = [
            value for value in defaults if value["sport"] != selected["sport"]
        ]
        updated.append(selected)
        updated.sort(key=lambda value: value["sport"])
        if len(updated) > MAX_PROFILE_DEFAULTS:
            raise LocalDraftProfileValidationError(
                f"local draft profile defaults cannot exceed {MAX_PROFILE_DEFAULTS} sports"
            )
        _write_profile_defaults(updated, defaults_destination)
    return selected


def clear_default_local_draft_profile(
    sport: str,
    path: str | Path | None = None,
) -> bool:
    """Atomically clear one sport's default without touching saved profiles."""

    clean_sport = _safe_string(sport, "sport", 32)
    if not _IDENTIFIER.fullmatch(clean_sport):
        raise LocalDraftProfileValidationError("sport has an invalid format")
    custom_path = path is not None or bool(
        os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_DEFAULTS_PATH")
    ) or bool(os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_PATH"))
    destination = _profile_defaults_store_path(path)
    with _STORE_LOCK:
        if not custom_path and destination.parent.is_symlink():
            raise LocalDraftProfileValidationError(
                "default local draft profile directory cannot be a symbolic link"
            )
        defaults = _read_profile_defaults(destination)
        updated = [value for value in defaults if value["sport"] != clean_sport]
        if len(updated) == len(defaults):
            if destination.exists():
                destination.chmod(0o600)
            return False
        _prepare_store_directory(destination, tighten_existing=not custom_path)
        _write_profile_defaults(updated, destination)
    return True


def save_local_draft_profile(value: Any, path: str | Path | None = None) -> dict[str, Any]:
    """Atomically store one exact-identity local draft profile."""

    profile = sanitize_local_draft_profile(value)
    custom_path = path is not None or bool(os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_PATH"))
    destination = _profile_store_path(path)
    with _STORE_LOCK:
        _prepare_store_directory(destination, tighten_existing=not custom_path)
        sessions = _read_all(destination)
        session_key = profile["draft"]["sessionKey"]
        existing = sessions.get(session_key)
        if existing is not None:
            if existing["draft"] != profile["draft"]:
                raise LocalDraftProfileValidationError(
                    "local draft profile identity does not match the saved session"
                )
            if profile == existing:
                destination.chmod(0o600)
                return existing
            if _parse_timestamp(profile["importedAt"]) <= _parse_timestamp(existing["importedAt"]):
                raise LocalDraftProfileValidationError(
                    "local draft profile import must be newer than the saved profile"
                )
        sessions[session_key] = profile
        _write_all(sessions, destination)
    return profile


def load_local_draft_profile(
    draft_identity: Mapping[str, Any], path: str | Path | None = None
) -> dict[str, Any] | None:
    """Load only the profile whose full draft identity exactly matches the caller."""

    identity = _sanitize_draft_identity(draft_identity)
    custom_path = path is not None or bool(os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_PATH"))
    destination = _profile_store_path(path)
    with _STORE_LOCK:
        if not custom_path and destination.parent.is_symlink():
            raise LocalDraftProfileValidationError(
                "default local draft profile directory cannot be a symbolic link"
            )
        profile = _read_all(destination).get(identity["sessionKey"])
    if profile is None or profile["draft"] != identity:
        return None
    return profile


def list_local_draft_profile_summaries(
    path: str | Path | None = None,
) -> list[dict[str, Any]]:
    """List privacy-minimal metadata for explicit local profile selection."""

    custom_path = path is not None or bool(os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_PATH"))
    destination = _profile_store_path(path)
    with _STORE_LOCK:
        if not custom_path and destination.parent.is_symlink():
            raise LocalDraftProfileValidationError(
                "default local draft profile directory cannot be a symbolic link"
            )
        profiles = list(_read_all(destination).values())
    summaries: list[dict[str, Any]] = []
    for profile in profiles:
        provenance = profile["provenance"]
        summary: dict[str, Any] = {
            "sport": profile["draft"]["sport"],
            "leagueId": profile["draft"]["leagueId"],
            "importedAt": profile["importedAt"],
            "format": provenance["format"],
            "rankingCount": len(profile["rankings"]),
        }
        if "asOf" in provenance:
            summary["asOf"] = provenance["asOf"]
        summaries.append(summary)
    return sorted(
        summaries,
        key=lambda summary: (
            summary["importedAt"],
            summary["sport"],
            summary["leagueId"],
        ),
        reverse=True,
    )


def _reusable_profile_content(profile: Mapping[str, Any]) -> dict[str, Any]:
    return {
        field: profile[field]
        for field in ("season", "rankings", "leagueSettings", "provenance")
    }


def bind_local_draft_profile(
    source_league_id: str,
    target_draft_identity: Mapping[str, Any],
    path: str | Path | None = None,
) -> dict[str, Any]:
    """Explicitly copy sanitized profile data onto one exact target draft identity."""

    target = _sanitize_draft_identity(target_draft_identity)
    source_id = _safe_string(source_league_id, "sourceLeagueId", 64)
    if not _IDENTIFIER.fullmatch(source_id):
        raise LocalDraftProfileValidationError("sourceLeagueId has an invalid format")
    custom_path = path is not None or bool(os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_PATH"))
    destination = _profile_store_path(path)
    with _STORE_LOCK:
        if not custom_path and destination.parent.is_symlink():
            raise LocalDraftProfileValidationError(
                "default local draft profile directory cannot be a symbolic link"
            )
        _prepare_store_directory(destination, tighten_existing=not custom_path)
        profiles = _read_all(destination)
        source_key = f"{target['sport']}:{source_id}"
        source = profiles.get(source_key)
        if source is None:
            source_sports = {
                profile["draft"]["sport"]
                for profile in profiles.values()
                if profile["draft"]["leagueId"] == source_id
            }
            if source_sports:
                raise LocalDraftProfileConflictError(
                    "selected local profile belongs to a different sport"
                )
            raise LocalDraftProfileNotFoundError("selected local draft profile was not found")

        existing = profiles.get(target["sessionKey"])
        if existing is not None:
            if existing["draft"] != target:
                raise LocalDraftProfileConflictError(
                    "target local profile identity does not match the synced draft"
                )
            if _reusable_profile_content(existing) == _reusable_profile_content(source):
                destination.chmod(0o600)
                return existing
            raise LocalDraftProfileConflictError(
                "selected draft already has a different local profile"
            )

        bound = deepcopy(source)
        bound["draft"] = target
        bound = sanitize_local_draft_profile(bound)
        profiles[target["sessionKey"]] = bound
        _write_all(profiles, destination)
    return bound


def bind_default_local_draft_profile(
    target_draft_identity: Mapping[str, Any],
    *,
    profile_path: str | Path | None = None,
    defaults_path: str | Path | None = None,
    current_season: int | None = None,
) -> dict[str, Any] | None:
    """Bind the selected same-sport default to an otherwise unbound draft."""

    target = _sanitize_draft_identity(target_draft_identity)
    profile_destination = _profile_store_path(profile_path)
    defaults_destination = _profile_defaults_store_path(
        defaults_path, profile_path=profile_path
    )
    custom_profile_path = profile_path is not None or bool(
        os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_PATH")
    )
    custom_defaults_path = (
        defaults_path is not None
        or bool(os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_DEFAULTS_PATH"))
        or profile_path is not None
        or bool(os.getenv("FANTASY_FOOTBALL_DRAFT_PROFILE_PATH"))
    )
    with _STORE_LOCK:
        if not custom_profile_path and profile_destination.parent.is_symlink():
            raise LocalDraftProfileValidationError(
                "default local draft profile directory cannot be a symbolic link"
            )
        profiles = _read_all(profile_destination)
        existing = profiles.get(target["sessionKey"])
        if existing is not None:
            if existing["draft"] != target:
                raise LocalDraftProfileConflictError(
                    "target local profile identity does not match the synced draft"
                )
            return existing

        if not custom_defaults_path and defaults_destination.parent.is_symlink():
            raise LocalDraftProfileValidationError(
                "default local draft profile directory cannot be a symbolic link"
            )
        defaults = _read_profile_defaults(defaults_destination)
        selected = next(
            (value for value in defaults if value["sport"] == target["sport"]),
            None,
        )
        if selected is None:
            return None
        source_key = f"{target['sport']}:{selected['sourceLeagueId']}"
        source = profiles.get(source_key)
        if source is None:
            source_sports = {
                profile["draft"]["sport"]
                for profile in profiles.values()
                if profile["draft"]["leagueId"] == selected["sourceLeagueId"]
            }
            if source_sports:
                raise LocalDraftProfileConflictError(
                    "default local profile belongs to a different sport"
                )
            raise LocalDraftProfileNotFoundError(
                "default local draft profile source was not found"
            )
        _require_current_default_profile_season(source, current_season)

        bound = deepcopy(source)
        bound["draft"] = target
        bound = sanitize_local_draft_profile(bound)
        _prepare_store_directory(
            profile_destination, tighten_existing=not custom_profile_path
        )
        profiles[target["sessionKey"]] = bound
        _write_all(profiles, profile_destination)
    return bound


def _header_key(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    normalized = unicodedata.normalize("NFKC", value).upper()
    return re.sub(r"[^A-Z0-9]+", "", normalized)


def _row_lookup(row: Mapping[str, Any], aliases: set[str]) -> Any:
    matches = [value for key, value in row.items() if _header_key(key) in aliases]
    if len(matches) > 1:
        raise LocalDraftProfileValidationError(
            "DraftSheets row contains ambiguous duplicate columns"
        )
    return matches[0] if matches else None


def _is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _is_optional_blank(value: Any) -> bool:
    return _is_blank(value) or (
        isinstance(value, str) and value.strip().upper() in {"-", "N/A", "#N/A"}
    )


def _convert_ecr_rows(ecr_rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    if isinstance(ecr_rows, (str, bytes)) or not isinstance(ecr_rows, Sequence):
        raise LocalDraftProfileValidationError("ECR rows must be a sequence")
    if len(ecr_rows) > MAX_ECR_ROWS:
        raise LocalDraftProfileValidationError(f"ECR rows cannot exceed {MAX_ECR_ROWS}")
    rankings: list[dict[str, Any]] = []
    for row_number, row in enumerate(ecr_rows, start=1):
        if not isinstance(row, Mapping):
            raise LocalDraftProfileValidationError(
                f"DraftSheets ECR row {row_number} must be an object"
            )
        rank_value = _row_lookup(row, _ECR_FIELD_ALIASES["rank"])
        name_value = _row_lookup(row, _ECR_FIELD_ALIASES["name"])
        team_value = _row_lookup(row, _ECR_FIELD_ALIASES["team"])
        position_value = _row_lookup(row, _ECR_FIELD_ALIASES["position"])
        required = (rank_value, name_value, team_value, position_value)
        if all(_is_blank(item) for item in required):
            continue
        if any(_is_blank(item) for item in required):
            raise LocalDraftProfileValidationError(
                f"DraftSheets ECR row {row_number} is incomplete"
            )
        candidate: dict[str, Any] = {
            "name": _safe_player_name(name_value, f"ECR row {row_number} player"),
            "position": _normalize_position(position_value, roster=False),
            "team": _safe_string(team_value, f"ECR row {row_number} team", 8).upper(),
            "rank": _coerce_integer(rank_value, f"ECR row {row_number} rank", 1, 10_000),
        }
        adp_value = _row_lookup(row, _ECR_FIELD_ALIASES["adp"])
        if not _is_optional_blank(adp_value):
            candidate["average_draft_position"] = _coerce_number(
                adp_value, f"ECR row {row_number} ADP", 0.01, 10_000
            )
        bye_value = _row_lookup(row, _ECR_FIELD_ALIASES["bye"])
        if not _is_optional_blank(bye_value):
            candidate["bye_week"] = _coerce_integer(
                bye_value, f"ECR row {row_number} bye week", 1, 22
            )
        player_key_value = _row_lookup(row, _ECR_FIELD_ALIASES["player_key"])
        if not _is_optional_blank(player_key_value):
            player_key = normalize_yahoo_player_key(player_key_value)
            if player_key is None:
                raise LocalDraftProfileValidationError(
                    f"ECR row {row_number} player_key has an invalid format"
                )
            candidate["player_key"] = player_key
        rankings.append(candidate)
    if not rankings:
        raise LocalDraftProfileValidationError(
            "DraftSheets ECR must contain at least one candidate"
        )
    rankings.sort(key=lambda candidate: (candidate["rank"], candidate["name"]))
    return rankings[:MAX_CANDIDATES]


def _scoring_name(value: Any) -> str | None:
    return _SCORING_ALIASES.get(_header_key(value))


def _merge_scoring_value(settings: dict[str, int], name: str, value: Any, row_number: int) -> None:
    if _is_blank(value):
        return
    maximum = 20 if name == "teams" else 30
    minimum = 2 if name == "teams" else 0
    parsed = _coerce_integer(value, f"Scoring row {row_number} {name}", minimum, maximum)
    if name != "teams" and parsed == 0:
        return
    existing = settings.get(name)
    if existing is not None and existing != parsed:
        raise LocalDraftProfileValidationError(f"DraftSheets Scoring has conflicting {name} values")
    settings[name] = parsed


def _convert_scoring_rows(
    scoring_rows: Sequence[Mapping[str, Any]] | Mapping[str, Any],
) -> dict[str, Any]:
    if isinstance(scoring_rows, Mapping):
        rows: list[Mapping[str, Any]] = [scoring_rows]
    elif isinstance(scoring_rows, Sequence) and not isinstance(scoring_rows, (str, bytes)):
        rows = list(scoring_rows)
    else:
        raise LocalDraftProfileValidationError("Scoring rows must be a sequence")
    if not rows or len(rows) > MAX_SCORING_ROWS:
        raise LocalDraftProfileValidationError(
            f"Scoring rows must contain between 1 and {MAX_SCORING_ROWS} rows"
        )
    if any(not isinstance(row, Mapping) for row in rows):
        raise LocalDraftProfileValidationError("each DraftSheets Scoring row must be an object")

    settings: dict[str, int] = {}
    grid_value_cells = {
        (index + 1, column)
        for index, row in enumerate(rows[:-1])
        for column, label in row.items()
        if _scoring_name(label) is not None and column in rows[index + 1]
    }
    for index, row in enumerate(rows):
        row_number = index + 1
        # Semantic column mappings, such as {"#TEAMS:": 12, "QB:": 1}.
        for key, value in row.items():
            name = _scoring_name(key)
            # A one-letter spreadsheet column such as K can also name a position.
            # A recognized label in its value means this is a grid header instead.
            if (
                name is not None
                and _scoring_name(value) is None
                and (index, key) not in grid_value_cells
            ):
                _merge_scoring_value(settings, name, value, row_number)

        # Vertical pair mappings, such as {"setting": "QB", "value": 1}.
        normalized_keys = {_header_key(key): value for key, value in row.items()}
        label = normalized_keys.get("SETTING")
        if label is not None and "VALUE" in normalized_keys:
            name = _scoring_name(label)
            if name is not None:
                _merge_scoring_value(settings, name, normalized_keys["VALUE"], row_number)

        # Spreadsheet grids, where one row contains labels and the next the values.
        if index + 1 < len(rows):
            next_row = rows[index + 1]
            for column, label in row.items():
                name = _scoring_name(label)
                if name is not None and column in next_row:
                    _merge_scoring_value(settings, name, next_row[column], row_number + 1)

    teams = settings.pop("teams", None)
    if teams is None:
        raise LocalDraftProfileValidationError("DraftSheets Scoring must include team count")
    positions = [
        {"position": position, "count": settings[position]}
        for position in _ROSTER_ORDER
        if position in settings
    ]
    if not positions:
        raise LocalDraftProfileValidationError("DraftSheets Scoring must include roster positions")
    return {"teams": teams, "rosterPositions": positions}


def profile_from_draftsheets_rows(
    ecr_rows: Sequence[Mapping[str, Any]],
    scoring_rows: Sequence[Mapping[str, Any]] | Mapping[str, Any],
    *,
    draft: Mapping[str, Any],
    imported_at: str,
    season: int,
    as_of: str | None = None,
) -> dict[str, Any]:
    """Convert pre-extracted DraftSheets rows into the canonical local profile.

    Only exact allowlisted semantic columns are consumed. In particular,
    ``ECR VS. ADP`` is a delta and is never mislabeled as player ADP.
    """

    provenance: dict[str, Any] = {
        "kind": "user-import",
        "format": "draftsheets-2026",
    }
    if as_of is not None:
        provenance["asOf"] = as_of
    return sanitize_local_draft_profile(
        {
            "schemaVersion": 1,
            "source": "local-draft-profile",
            "season": season,
            "importedAt": imported_at,
            "draft": dict(draft),
            "rankings": _convert_ecr_rows(ecr_rows),
            "leagueSettings": _convert_scoring_rows(scoring_rows),
            "provenance": provenance,
        }
    )


def local_draft_profile_revision(value: Any) -> str:
    """Return a stable content revision for post-scoring change detection."""

    profile = sanitize_local_draft_profile(value)
    canonical = json.dumps(
        profile,
        ensure_ascii=False,
        allow_nan=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")
    return hashlib.sha256(canonical).hexdigest()


def _validate_xlsx_archive(workbook: bytes) -> None:
    if not isinstance(workbook, bytes) or not workbook:
        raise LocalDraftProfileValidationError("DraftSheets workbook must be non-empty bytes")
    if len(workbook) > MAX_XLSX_BYTES:
        raise LocalDraftProfileValidationError(
            f"DraftSheets workbook cannot exceed {MAX_XLSX_BYTES} bytes"
        )
    try:
        with zipfile.ZipFile(io.BytesIO(workbook)) as archive:
            members = archive.infolist()
            if len(members) > MAX_XLSX_MEMBERS:
                raise LocalDraftProfileValidationError(
                    f"DraftSheets workbook cannot exceed {MAX_XLSX_MEMBERS} ZIP members"
                )
            total_size = 0
            seen_parts: set[str] = set()
            for member in members:
                path = PurePosixPath(member.filename)
                parts = path.parts
                if path.is_absolute() or ".." in parts or "\\" in member.filename:
                    raise LocalDraftProfileValidationError(
                        "DraftSheets workbook contains an unsafe ZIP path"
                    )
                lowered = member.filename.casefold()
                if lowered in seen_parts:
                    raise LocalDraftProfileValidationError(
                        "DraftSheets workbook contains duplicate ZIP parts"
                    )
                seen_parts.add(lowered)
                if member.flag_bits & 0x1:
                    raise LocalDraftProfileValidationError(
                        "encrypted DraftSheets workbook parts are not accepted"
                    )
                total_size += member.file_size
                if total_size > MAX_XLSX_UNCOMPRESSED_BYTES:
                    raise LocalDraftProfileValidationError(
                        "DraftSheets workbook expands beyond 16 MiB"
                    )
                if (
                    lowered.endswith(".bin")
                    or "vbaproject" in lowered
                    or lowered.startswith(_MACRO_PART_PREFIXES)
                ):
                    raise LocalDraftProfileValidationError(
                        "DraftSheets workbooks with macros are not accepted"
                    )
                if lowered.startswith("xl/externallinks/"):
                    raise LocalDraftProfileValidationError(
                        "DraftSheets workbooks with external links are not accepted"
                    )
                xml_content: bytes | None = None
                if lowered.endswith((".xml", ".rels")):
                    xml_content = archive.read(member).lower()
                    if (
                        b"\x00" in xml_content
                        or b"<!doctype" in xml_content
                        or b"<!entity" in xml_content
                    ):
                        raise LocalDraftProfileValidationError(
                            "DraftSheets workbook contains unsafe XML"
                        )
                if lowered.endswith(".rels") or lowered == "[content_types].xml":
                    package_metadata = xml_content or b""
                    if any(marker in package_metadata for marker in _MACRO_PACKAGE_MARKERS):
                        raise LocalDraftProfileValidationError(
                            "DraftSheets workbooks with macros are not accepted"
                        )
                    # Internal OOXML relationships omit TargetMode. Rejecting the
                    # attribute entirely fails closed for hyperlinks, attached
                    # templates, external workbooks, and other remote targets without
                    # retaining or echoing their URLs.
                    if lowered.endswith(".rels") and (
                        b"targetmode" in package_metadata
                        or b"relationships/externallink" in package_metadata
                    ):
                        raise LocalDraftProfileValidationError(
                            "DraftSheets workbooks with external links are not accepted"
                        )
    except zipfile.BadZipFile as error:
        raise LocalDraftProfileValidationError(
            "DraftSheets workbook is not a valid XLSX archive"
        ) from error


def _workbook_as_of(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, str):
        try:
            return date.fromisoformat(value.strip()).isoformat()
        except ValueError:
            return None
    return None


def profile_from_draftsheets_xlsx(
    workbook: bytes,
    *,
    draft: Mapping[str, Any],
    imported_at: str,
    season: int,
    as_of: str | None = None,
    roster_overrides: Mapping[str, Any] | None = None,
) -> dict[str, Any]:
    """Read a bounded DraftSheets workbook and return a sanitized local profile."""

    _validate_xlsx_archive(workbook)
    try:
        from openpyxl import load_workbook
    except ImportError as error:  # pragma: no cover - dependency is declared by the app
        raise LocalDraftProfileValidationError(
            "DraftSheets XLSX import requires openpyxl"
        ) from error
    try:
        parsed = load_workbook(
            io.BytesIO(workbook), read_only=True, data_only=True, keep_links=False
        )
    except Exception as error:
        raise LocalDraftProfileValidationError(
            "DraftSheets workbook could not be parsed"
        ) from error
    try:
        if "ECR" not in parsed.sheetnames or "Scoring" not in parsed.sheetnames:
            raise LocalDraftProfileValidationError(
                "DraftSheets workbook must include ECR and Scoring sheets"
            )
        ecr_sheet = parsed["ECR"]
        header = next(
            ecr_sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=10, values_only=True),
            (),
        )
        expected = {
            0: "RK",
            2: "PLAYERNAME",
            3: "TEAM",
            4: "POS",
            5: "BYEWEEK",
        }
        if any(
            index >= len(header) or _header_key(header[index]) != expected_name
            for index, expected_name in expected.items()
        ):
            raise LocalDraftProfileValidationError(
                "DraftSheets ECR columns do not match the supported 2026 format"
            )
        ecr_rows: list[dict[str, Any]] = []
        for row in ecr_sheet.iter_rows(
            min_row=2,
            max_row=MAX_ECR_ROWS + 1,
            min_col=1,
            max_col=10,
            values_only=True,
        ):
            extracted = {
                "RK": row[0],
                "PLAYER NAME": row[2],
                "TEAM": row[3],
                "POS": row[4],
                "BYE WEEK": row[5],
            }
            ecr_rows.append(extracted)
            if (
                len(ecr_rows) >= MAX_CANDIDATES
                and sum(not all(_is_blank(value) for value in item.values()) for item in ecr_rows)
                >= MAX_CANDIDATES
            ):
                break

        scoring_sheet = parsed["Scoring"]
        header_row = {
            str(column): scoring_sheet.cell(row=3, column=column).value for column in range(9, 27)
        }
        value_row = {
            str(column): scoring_sheet.cell(row=4, column=column).value for column in range(9, 27)
        }
        scoring_rows: list[Mapping[str, Any]] = [header_row, value_row]
        resolved_as_of = as_of or _workbook_as_of(scoring_sheet["B1"].value)
        profile = profile_from_draftsheets_rows(
            ecr_rows,
            scoring_rows,
            draft=draft,
            imported_at=imported_at,
            season=season,
            as_of=resolved_as_of,
        )
        if roster_overrides is not None:
            if not isinstance(roster_overrides, Mapping) or not roster_overrides:
                raise LocalDraftProfileValidationError("roster overrides must be an object")
            positions = {
                entry["position"]: entry["count"]
                for entry in profile["leagueSettings"]["rosterPositions"]
            }
            overridden: set[str] = set()
            for raw_position, raw_count in roster_overrides.items():
                position = _normalize_position(raw_position, roster=True)
                if position in overridden:
                    raise LocalDraftProfileValidationError(
                        f"duplicate roster override: {position}"
                    )
                overridden.add(position)
                count = _coerce_integer(
                    raw_count,
                    f"roster override {position}",
                    0,
                    30,
                )
                if count == 0:
                    positions.pop(position, None)
                else:
                    positions[position] = count
            profile["leagueSettings"] = _sanitize_league_settings(
                {
                    "teams": profile["leagueSettings"]["teams"],
                    "rosterPositions": [
                        {"position": position, "count": positions[position]}
                        for position in _ROSTER_ORDER
                        if position in positions
                    ],
                }
            )
            profile = sanitize_local_draft_profile(profile)
        return profile
    finally:
        parsed.close()
